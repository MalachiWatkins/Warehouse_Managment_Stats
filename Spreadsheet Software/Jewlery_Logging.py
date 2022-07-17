from pymongo import MongoClient
import pymongo
import openpyxl
from openpyxl import load_workbook
import time
from plyer import notification

warehouse_db = cluster["WAREHOUSE_MANAGEMENT"]
FINISHEDCollection = warehouse_db["FINISHED"]
Finished_JewlCollection = warehouse_db["FINISHED_JEWL"]


def gen(dnd, spreadsheet, collection, type):
    with open(dnd, "r") as collum:
        empty_collum = collum.readline()
    wb = load_workbook(spreadsheet)
    sheet = wb.active

    # y = empty_collum
    x = 0

    document_list = ['Storage_Type', 'Date_Received', 'Store_Number',
                     'Contents', 'Date_Processed', 'MANIFEST_NUMBER', 'Problems']
    single_doc_list = []
    for document in collection.find():
        single_doc_list.append(document)
    if single_doc_list == []:
        notification.notify(
            title='No Processed Jewlery',
            message='Nothing was processed right now!',
            timeout=10,
        )
        exit()
    else:
        notification.notify(
            title='Adding Processed Jewelry to Spreadsheet',
            message='The Process Has Started!',
            timeout=5,
        )
        while x < len(single_doc_list):
            empty_collum = int(empty_collum) + 1
            print(single_doc_list[x]['Storage_Type'])
            Date_Received = sheet.cell(row=int(empty_collum), column=1)
            Date_Received.value = str(single_doc_list[x]['Date_Received'])

            store_number = sheet.cell(row=int(empty_collum), column=2)
            store_number.value = str(single_doc_list[x]['Store_Number'])

            processed_by = sheet.cell(row=int(empty_collum), column=3)
            processed_by.value = str(single_doc_list[x]['Processed_By'])

            Date_Processed = sheet.cell(row=int(empty_collum), column=4)
            Date_Processed.value = str(single_doc_list[x]['Date_Processed'])

            MANIFEST_NUMBER = sheet.cell(row=int(empty_collum), column=5)
            MANIFEST_NUMBER.value = str(single_doc_list[x]['MANIFEST_NUMBER'])

            Problems = sheet.cell(row=int(empty_collum), column=6)
            Problems.value = str(single_doc_list[x]['Problems'])

            seal_number = sheet.cell(row=int(empty_collum), column=7)
            seal_number.value = str(single_doc_list[x]['Seal_Number'])

            # delquery = { "_id": float(single_doc_list[x]['_id']) }
            # collection.delete_one(delquery)

            x += 1
        with open(dnd, "w") as coll:
            empty = empty_collum
            coll.write(str(empty))
        notification.notify(
            title='Finished Adding Items to Spreadsheet',
            message='All Done Here :)',
            timeout=5,
        )
        wb.save(filename=spreadsheet)
    return


gen(dnd='DO_NOT_DELETE_JEWL.txt', spreadsheet='Jewlery_Log.xlsx',
    collection=Finished_JewlCollection, type='jewlery')
time.sleep(100)
