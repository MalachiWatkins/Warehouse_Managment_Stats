from pymongo import MongoClient
import pymongo
import openpyxl
from openpyxl import load_workbook
import time
from plyer import notification
import random
# THIS IS ALL PROCESSED DATA

warehouse_db = cluster["WAREHOUSE_MANAGEMENT_GOODWILL"]
Truck_Receiver_DB = warehouse_db["Truck_Receiver_DB"]

Processor_Review_DB = warehouse_db["Processor_Review_DB"]


Jewelry_DB = warehouse_db["Jewelry_DB"]
Jewelry_Review_DB = warehouse_db["Jewelry_Review_DB"]

Books_Media_DB = warehouse_db["Books_Media_DB"]
Book_Media_Review_DB = warehouse_db["Book_Media_Review_DB"]

Finished_DB = warehouse_db["Finished_DB"]

Finished_Jewlery_DB = warehouse_db["Finished_Jewlery_DB"]

ARCHIVE_COLLECTION = warehouse_db["ARCHIVE"]


def gen(dnd, spreadsheet, collection):
    # Opens WorkBook
    with open(dnd, "r") as collum:
        empty_collum = collum.readline()
    wb = load_workbook(spreadsheet)
    sheet = wb.active

    # y = empty_collum
    x = 0
    single_doc_list = []
    for document in collection.find():
        single_doc_list.append(document)
        document['_id'] = random.random()
        ARCHIVE_COLLECTION.insert_one(document)
        #ID = document['_id']
        #delquery = {"_id": float(ID)}
        # collection.delete_one(delquery)
    print(len(single_doc_list))
    if single_doc_list == []:
        notification.notify(
            title='No Processed Items',
            message='Nothing was processed right now!',
            timeout=100,
        )
        exit()
    else:
        notification.notify(
            title='Adding Processed Itmes to Spreadsheet',
            message='The Process Has Started!',
            timeout=100,
        )
        while x < len(single_doc_list):

            empty_collum = int(empty_collum) + 1

            try:
                Processed_By = sheet.cell(row=int(empty_collum), column=1)
                Processed_By.value = str(single_doc_list[x]['Processed_By'])
            except:
                Processed_By = sheet.cell(row=int(empty_collum), column=1)
                Processed_By.value = str('N/A')

            storage_type = sheet.cell(row=int(empty_collum), column=2)
            storage_type.value = str(single_doc_list[x]['Storage_Type'])

            Date_Received = sheet.cell(row=int(empty_collum), column=3)
            Date_Received.value = str(single_doc_list[x]['Date_Received'])

            store_number = sheet.cell(row=int(empty_collum), column=4)
            store_number.value = str(single_doc_list[x]['Store_Number'])

            Contents = sheet.cell(row=int(empty_collum), column=5)
            Contents.value = str(single_doc_list[x]['Contents'])

            Date_Processed = sheet.cell(row=int(empty_collum), column=6)
            Date_Processed.value = str(single_doc_list[x]['Date_Processed'])
            if str(single_doc_list[x]['MANIFEST_NUMBER']) == '':
                MANIFEST_NUMBER = sheet.cell(row=int(empty_collum), column=7)
                MANIFEST_NUMBER.value = str("No Manifest")
            else:
                MANIFEST_NUMBER = sheet.cell(row=int(empty_collum), column=7)
                MANIFEST_NUMBER.value = str(
                    single_doc_list[x]['MANIFEST_NUMBER'])
            try:
                seal_number = sheet.cell(row=int(empty_collum), column=8)
                seal_number.value = str(single_doc_list[x]['Seal_Number'])
            except:
                seal_number = sheet.cell(row=int(empty_collum), column=8)
                seal_number.value = str('N/A')
            if str(single_doc_list[x]['Problems']) == '':
                Problems = sheet.cell(row=int(empty_collum), column=9)
                Problems.value = str('N/A')
            else:
                Problems = sheet.cell(row=int(empty_collum), column=9)
                Problems.value = str(single_doc_list[x]['Problems'])

            x += 1
        with open(dnd, "w") as coll:
            empty = empty_collum
            coll.write(str(empty))
        notification.notify(
            title='Finished Adding Items to Spreadsheet',
            message='All Done Here :)',
            timeout=5,
        )
        wb.save(filename=spreadsheet)
    return


##############################################################
########## AUTO GEN GOES HERE 7am and 4pm EST ###############
##############################################################
gen(dnd='DO_NOT_DELETE_SGW.txt', spreadsheet='Gay_lord_Toat_Log.xlsx',
    collection=Finished_DB)
gen(dnd='DO_NOT_DELETE_JEWL.txt', spreadsheet='Jewlery_Log.xlsx',
    collection=Finished_Jewlery_DB)
# ADD GEN for empty spreasheet
LIST = [Truck_Receiver_DB, Processor_Review_DB, Jewelry_DB,
        Jewelry_Review_DB, Books_Media_DB, Book_Media_Review_DB]
for COLLECTION in LIST:
    gen(dnd='DO_NOT_DELETE_UNPROCESSED.txt', spreadsheet='Unprocessed.xlsx',
        collection=COLLECTION)

time.sleep(100)
