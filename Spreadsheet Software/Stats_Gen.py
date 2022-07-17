from pymongo import MongoClient
import pymongo
import openpyxl
from openpyxl import load_workbook
import time
from plyer import notification
import random
# THIS IS ALL PROCESSED DATA

warehouse_db = cluster["WAREHOUSE_MANAGEMENT_GOODWILL"]
Truck_Receiver_DB = warehouse_db["Truck_Receiver_DB"]

Processor_Review_DB = warehouse_db["Processor_Review_DB"]


Jewelry_DB = warehouse_db["Jewelry_DB"]
Jewelry_Review_DB = warehouse_db["Jewelry_Review_DB"]

Books_Media_DB = warehouse_db["Books_Media_DB"]
Book_Media_Review_DB = warehouse_db["Book_Media_Review_DB"]

Finished_DB = warehouse_db["Finished_DB"]

Finished_Jewlery_DB = warehouse_db["Finished_Jewlery_DB"]

ARCHIVE_COLLECTION = warehouse_db["ARCHIVE"]


def gen(dnd, spreadsheet, collection, unprocessed):
    # Opens WorkBook
    with open(dnd, "r") as collum:
        empty_collum = collum.readline()
    wb = load_workbook(spreadsheet)
    sheet = wb.active

    # y = empty_collum
    x = 0
    single_doc_list = []
    for document in collection.find():
        single_doc_list.append(document)
        document['_id'] = random.random()
        ARCHIVE_COLLECTION.insert_one(document)
        #ID = document['_id']
        #delquery = {"_id": float(ID)}
        # collection.delete_one(delquery)
    print(len(single_doc_list))
    if single_doc_list == []:
        exit()
    else:

        while x < len(single_doc_list):

            empty_collum = int(empty_collum) + 1

            try:
                Processed_By = sheet.cell(row=int(empty_collum), column=1)
                Processed_By.value = str(single_doc_list[x]['Processed_By'])
            except:
                Processed_By = sheet.cell(row=int(empty_collum), column=1)
                Processed_By.value = str('N/A')

            storage_type = sheet.cell(row=int(empty_collum), column=2)
            storage_type.value = str(single_doc_list[x]['Storage_Type'])

            Date_Received = sheet.cell(row=int(empty_collum), column=3)
            Date_Received.value = str(single_doc_list[x]['Date_Received'])

            store_number = sheet.cell(row=int(empty_collum), column=4)
            store_number.value = str(single_doc_list[x]['Store_Number'])

            Contents = sheet.cell(row=int(empty_collum), column=5)
            Contents.value = str(single_doc_list[x]['Contents'])
            if unprocessed == True:
                Date_Processed = sheet.cell(row=int(empty_collum), column=6)
                Date_Processed.value = str('Unporcessed')
            else:
                Date_Processed = sheet.cell(row=int(empty_collum), column=6)
                Date_Processed.value = str(single_doc_list[x]['Date_Processed'])
            try:
                if str(single_doc_list[x]['MANIFEST_NUMBER']) == '':
                    MANIFEST_NUMBER = sheet.cell(row=int(empty_collum), column=7)
                    MANIFEST_NUMBER.value = str("No Manifest")
                else:
                    MANIFEST_NUMBER = sheet.cell(row=int(empty_collum), column=7)
                    MANIFEST_NUMBER.value = str(
                        single_doc_list[x]['MANIFEST_NUMBER'])
            except:
                MANIFEST_NUMBER = sheet.cell(row=int(empty_collum), column=7)
                MANIFEST_NUMBER.value = str("No Manifest")


            try:
                seal_number = sheet.cell(row=int(empty_collum), column=8)
                seal_number.value = str(single_doc_list[x]['Seal_Number'])
            except:
                seal_number = sheet.cell(row=int(empty_collum), column=8)
                seal_number.value = str('N/A')
            if str(single_doc_list[x]['Problems']) == '':
                Problems = sheet.cell(row=int(empty_collum), column=9)
                Problems.value = str('N/A')
            else:
                Problems = sheet.cell(row=int(empty_collum), column=9)
                Problems.value = str(single_doc_list[x]['Problems'])

            x += 1
        with open(dnd, "w") as coll:
            empty = empty_collum
            coll.write(str(empty))
        wb.save(filename=spreadsheet)
    return
def stats():
    book_query = {"Contents": 'Books'}
    media_query = {"Contents": "Media"}
    sgw_query = {"Contents": "Collectables"}
    jewlery_query = {"Contents": "Jewelry"}


    total_not_added_to_spreadsheet_books = Finished_DB.count_documents({
        "Contents": "Books"})
    total_not_added_to_spreadsheet_media = Finished_DB.count_documents({
        "Contents": "Media"})
    total_not_added_to_spreadsheet_sgw = Finished_DB.count_documents(
        {"Contents": "Collectables"})
    total_not_added_to_spreadsheet_jewlery = Finished_Jewlery_DB.count_documents({
        "Contents": "Jewelry"})
    # Total Left to process
    total_books_gay = Books_Media_DB.count_documents(
        {"Contents": "Books", "Storage_Type": "Gaylord"})
    total_books_tote = Books_Media_DB.count_documents(
        {"Contents": "Books", "Storage_Type": "Tote"})
    total_media_gay = Books_Media_DB.count_documents(
        {"Contents": "Media", "Storage_Type": "Gaylord"})
    total_media_tote = Books_Media_DB.count_documents(
        {"Contents": "Media", "Storage_Type": "Tote"})
    total_sgw_gay = Truck_Receiver_DB.count_documents(
        {"Contents": "Collectables", "Storage_Type": "Gaylord"})
    total_sgw_tote = Truck_Receiver_DB.count_documents(
        {"Contents": "Collectables", "Storage_Type": "Tote"})
    total_jewlery_gay = Jewelry_DB.count_documents(
        {"Contents": "Jewelry", "Storage_Type": "Gaylord"})
    total_jewlery_tote = Jewelry_DB.count_documents(
        {"Contents": "Jewelry", "Storage_Type": "Tote"})
    # Processed But not submitted

    Total_to_process_gay = total_books_gay + \
        total_media_gay + total_sgw_gay + total_jewlery_gay
    Total_to_process_tote = total_books_tote + \
        total_media_tote + total_sgw_tote + total_jewlery_tote

    total_Processed = total_not_added_to_spreadsheet_books + total_not_added_to_spreadsheet_media + \
        total_not_added_to_spreadsheet_sgw + total_not_added_to_spreadsheet_jewlery
    wb = load_workbook('Today.xlsx')
    sheet = wb.active
    ########### Total (SGW+JEWL+Books+Media) ################
    PROCESSED = sheet.cell(row=2, column=2)
    PROCESSED.value = str(total_Processed)

    total_Processed_SGW = sheet.cell(row=3, column=2)
    total_Processed_SGW.value = str(total_not_added_to_spreadsheet_sgw)

    total_Processed_Books = sheet.cell(row=4, column=2)
    total_Processed_Books.value = str(total_not_added_to_spreadsheet_books)

    total_Processed_Media = sheet.cell(row=5, column=2)
    total_Processed_Media.value = str(total_not_added_to_spreadsheet_media)

    total_Processed_Jewlery = sheet.cell(row=6, column=2)
    total_Processed_Jewlery.value = str(total_not_added_to_spreadsheet_jewlery)
    #
    TTPG = sheet.cell(row=9, column=2)
    TTPG.value = str(Total_to_process_gay)

    TTPT = sheet.cell(row=10, column=2)
    TTPT.value = str(Total_to_process_tote)
    #

    Total_SGW_Gay = sheet.cell(row=12, column=2)
    Total_SGW_Gay.value = str(total_sgw_gay)

    Total_SGW_Tote = sheet.cell(row=13, column=2)
    Total_SGW_Tote.value = str(total_sgw_tote)

    Total_Books_Gay = sheet.cell(row=15, column=2)
    Total_Books_Gay.value = str(total_books_gay)

    Total_Books_Tote = sheet.cell(row=16, column=2)
    Total_Books_Tote.value = str(total_books_tote)


    Total_Media_Gay = sheet.cell(row=18, column=2)
    Total_Media_Gay.value = str(total_media_gay)

    Total_Media_Tote = sheet.cell(row=19, column=2)
    Total_Media_Tote.value = str(total_media_tote)

    Total_Jewl_Gay = sheet.cell(row=21, column=2)
    Total_Jewl_Gay.value = str(total_jewlery_gay)

    Total_Jewl_Tote = sheet.cell(row=22, column=2)
    Total_Jewl_Tote.value = str(total_jewlery_tote)





    wb.save(filename='Today.xlsx')
    return

##############################################################
########## AUTO GEN GOES HERE 7am and 4pm EST ###############
##############################################################
stats()
gen(dnd='DO_NOT_DELETE_SGW.txt', spreadsheet='Gay_lord_Toat_Log.xlsx',
    collection=Finished_DB,unprocessed=False)
gen(dnd='DO_NOT_DELETE_JEWL.txt', spreadsheet='Jewlery_Log.xlsx',
    collection=Finished_Jewlery_DB, unprocessed=False)
# ADD GEN for empty spreasheet
LIST = [Truck_Receiver_DB, Processor_Review_DB, Jewelry_DB,
        Jewelry_Review_DB, Books_Media_DB, Book_Media_Review_DB]
for COLLECTION in LIST:
    gen(dnd='DO_NOT_DELETE_UNPROCESSED.txt', spreadsheet='Unporcessed.xlsx',
        collection=COLLECTION,unprocessed= True)

time.sleep(100)
